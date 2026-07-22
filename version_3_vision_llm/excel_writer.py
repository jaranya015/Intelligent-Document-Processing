"""
excel_writer.py

รับ dict ที่ได้จาก extract_receipt_data() แล้วเพิ่มเป็นแถวใหม่ในไฟล์ Excel
(สร้างไฟล์ + หัวตารางให้อัตโนมัติถ้ายังไม่มี)

ใช้ฟังก์ชันนี้ได้ทั้งตอนรัน batch กับรูปเก่า และตอนต่อกับ LINE bot ในอนาคต
"""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

HEADERS = [
    "timestamp",
    "source_file",
    "company",
    "document_type",
    "document_no",
    "date",
    "tax_id",
    "items_summary",
    "subtotal",
    "discount",
    "vat",
    "total",
    "currency",
    "low_confidence_fields",
]


def _summarize_items(items) -> str:
    if not items:
        return ""
    parts = []
    for it in items:
        name = it.get("name", "")
        qty = it.get("qty", "")
        amount = it.get("amount", "")
        parts.append(f"{name} x{qty} = {amount}")
    return "; ".join(parts)


def _ensure_workbook(excel_path: str):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Receipts"
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
    return wb, ws


def append_receipt_to_excel(data: dict, excel_path: str, source_file: str = "") -> None:
    """เพิ่มข้อมูลใบเสร็จ 1 รายการเป็นแถวใหม่ต่อท้ายไฟล์ Excel"""
    wb, ws = _ensure_workbook(excel_path)

    row = [
        datetime.now().isoformat(timespec="seconds"),
        source_file,
        data.get("company"),
        data.get("document_type"),
        data.get("document_no"),
        data.get("date"),
        data.get("tax_id"),
        _summarize_items(data.get("items")),
        data.get("subtotal"),
        data.get("discount"),
        data.get("vat"),
        data.get("total"),
        data.get("currency"),
        ", ".join(data.get("low_confidence_fields") or []),
    ]
    ws.append(row)
    wb.save(excel_path)